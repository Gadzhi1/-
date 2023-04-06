import os
import subprocess
from datetime import datetime, timedelta
import keyboard
import openpyxl
# from openpyxl.styles import PatternFill
import pandas as pd
import re
from past.builtins import raw_input
import cryptocode
import numpy as np
import matplotlib.pyplot as plt1
import telebot
from telebot.apihelper import ApiTelegramException
from keyboard import is_pressed
import win32com.client as win32

path_excel = 'долги.xlsx'
path_csv = "C:/Pfiles/loans_csv.csv"
path_subs_csv = "C:/Pfiles/loans_csv_subs.csv"
path_ord_csv = "C:/Pfiles/orders_csv.csv"
path_orc_csv = "C:/Pfiles/orders_supp_csv.csv"
path_ord_excel = 'заказы.xlsx'
path_orc_excel = 'пополнение товара.xlsx'
path_friends_csv = "C:/Pfiles/friends_csv.csv"
path_passwd = "C:/Pfiles/password.txt"
enc = 'UTF-8'

doc_loans = r'C:/Users/user190717/Desktop/долги.xlsx'
doc_orders = r'C:/Users/user190717/Desktop/заказы.xlsx'
doc_refills = r'C:/Users/user190717/Desktop/пополнение товара.xlsx'
sheet_loans = 'ДОЛГИ КЛИЕНТОВ'
sheet_orders = 'ЗАКАЗЫ ОТ КЛИЕНТОВ'
sheet_refills = 'ПОПОЛНЕНИЕ ТОВАРА'


def start_bot():
    os.startfile(r'C:/Users/user190717/Desktop/tm.exe')


def end_bot():
    subprocess.call(["taskkill", "/F", "/IM", "tm.exe"])


def end_app():
    subprocess.call(["taskkill", "/F", "/IM", "xh.exe"])


def format_excel(file1, sheet1):
    wb = openpyxl.load_workbook(file1)
    ws = wb[sheet1]

    for row in range(2, ws.max_row + 1):
        ws["{}{}".format("D", row)].number_format = '"$"#,##0_);[Red]("$"#,##0)'
    wb.save(file1)


def adjust_excel(file, sheet):
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file)
    ws = wb.Worksheets(sheet)
    ws.Columns.AutoFit()
    wb.Save()
    excel.Application.Quit()


def all_orc_func():
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    df_orc = pd.read_csv(path_orc_csv, engine='python', encoding=enc)
    df_orc.columns = ['товар', 'количество', 'дата']
    print(df_orc)
    print('\n')


def del_orc_func():
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    df_orc = pd.read_csv(path_orc_csv, engine='python', encoding=enc)
    df_orc.columns = ['товар', 'количество', 'дата']
    while True:
        print(df_orc.tail(10))
        print('\n')
        try:
            i = int(input(' ВВЕДИТЕ индекс: '))

            if i == 0:
                break

            if i not in df_orc.index:
                raise IndexError

            df_orc.drop(df_orc[df_orc.index == i].index, inplace=True)
            df_orc.to_csv(path_orc_csv, index=False)
            df_orc.to_excel(path_orc_excel, sheet_name="ПОПОЛНЕНИЕ ТОВАРА")
            adjust_excel(doc_refills, sheet_refills)
            print('\n' + '\t' + ' <<<<< Удалено >>>>>' + '\n')
        except ValueError:
            print('\n' + ' Ошибка! Индекс должен состоять из цифр!' + '\n')
        except IndexError:
            print('\n' + ' Ошибка! Такого индекса не существует!' + '\n')
        except PermissionError:
            print('\n' + ' Ошибка! Закройте файл!' + '\n')


def orc_func():
    while True:
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        df_orc = pd.read_csv(path_orc_csv, engine='python', encoding=enc)
        df_orc.columns = ['товар', 'количество', 'дата']
        print('\n')
        print(df_orc.tail(10))
        print('\n')
        while True:
            descr = input(' ВВЕДИТЕ товар: ').lower() or '- - -'

            if descr == '0' or descr == 'выйти' or descr == 'назад':
                break
            if len(descr) > 75:
                print(' Ошибка! Описание не должно превышать длину в 75 символов')
                continue
            break
        if descr == '0' or descr == 'выйти' or descr == 'назад':
            break

        while True:
            try:
                quantity = int(input(' ВВЕДИТЕ количество: '))
                if quantity == 0:
                    print('\n')
                    break
                if str(quantity)[:1] == '0':
                    raise ValueError(' Количество не должно начинаться с нуля')
                quantity = int(quantity)
            except ValueError:
                quantity = str('- - -')
                break
            else:
                break

        day = datetime.today().strftime('%d.%m.%Y')
        fcv = open(path_orc_csv, 'a', encoding=enc)
        fcv.write('{0},{1},{2}\n'.format(descr, quantity, day))
        fcv.close()

        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        df_orc = pd.read_csv(path_orc_csv, engine='python', encoding=enc)
        df_orc.columns = ['товар', 'количество', 'дата']

        df_orc.to_excel(path_orc_excel, sheet_name="ПОПОЛНЕНИЕ ТОВАРА")
        adjust_excel(doc_refills, sheet_refills)


def ord_func():
    while True:
        try:
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
            df_friends.columns = ['имя', 'телефон']

            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
            df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']

            day = datetime.today().strftime('%d.%m.%Y')

            if is_pressed('q'):
                print(' <--')
                break
            print('\n' + '\t' + '<<====================| ЗАКАЗЫ |====================>>' + '\n')
            print('        \'д\' ДОБАВИТЬ\n'
                  '        \'п\' ПОИСК\n'
                  '        \'т\' ТАБЛИЦА ЗАКАЗОВ\n'
                  '        \'и\' ИЗМЕНИТЬ\n'
                  '        \'у\' УДАЛИТЬ\n'
                  '        \'й + з\' НАЗАД')

            while True:
                if is_pressed('q'):
                    break
                if is_pressed('b'):
                    while True:
                        df_ord_copy = df_ord.copy()
                        df_ord_copy.drop('msg_id', axis=1, inplace=True)
                        print('\n')
                        print(df_ord_copy.tail(15))
                        print('\n')

                        i = int(input(' ВВЕДИТЕ индекс: '))

                        '''ed_row = df_ord.loc[df_ord.index == i]'''

                        choice_ed = int(input('\n'
                                              '        \'1\' ИМЯ\n'
                                              '        \'2\' КОММЕНТ\n'
                                              '        \'3\' ЗАЛОГ\n'
                                              '        \'4\' ТЕЛЕФОН\n'
                                              '        \'5\' ОПИСАНИЕ ЗАКАЗА\n'
                                              '    --> '))
                        if choice_ed == 0:
                            print(' ')
                            break

                        elif choice_ed == 1:
                            while True:

                                name = input('\n' + ' ВВЕДИТЕ новое имя: ')

                                if name == '0' or name == 'выйти' or name == 'назад':
                                    break

                                name = name[:1].upper() + name[1:].lower()

                                if not name.isalpha():
                                    print(' Ошибка! Имя не должно содержать цифры и спец. символы')
                                    continue

                                df_ord.iloc[[i], [0]] = name

                                sub_row = df_ord.loc[df_ord.index == i]

                                q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                   sub_row.iloc[0]['коммент'], \
                                                   sub_row.iloc[0]['телефон'], \
                                                   sub_row.iloc[0]['дата'], \
                                                   sub_row.iloc[0]['залог'], \
                                                   sub_row.iloc[0]['описание заказа']

                                message = f'#заказ\n' \
                                          f'Имя: #{q}\n' \
                                          f'Инф: {w}\n' \
                                          f'Тел: +7{e}\n' \
                                          f'Дата: {r}\n' \
                                          f'Залог: {t} руб.\n' \
                                          f'--> ( {y} )'

                                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                chatID = '-1001820758497'
                                bot = telebot.TeleBot(token)
                                bot.edit_message_text(text=message, chat_id=chatID,
                                                      message_id=sub_row.iloc[0]['msg_id'])

                                print('\n' + '\t' + '<<<<<< Имя изменено! >>>>>>' + '\n')
                                break

                            if name == '0' or name == 'выйти' or name == 'назад':
                                continue

                        elif choice_ed == 2:
                            comment = input('\n' + ' ВВЕДИТЕ новый коммент: ').lower() or '- - -'

                            if comment == '0' or comment == 'выйти' or comment == 'назад':
                                continue
                            if len(comment) > 15:
                                print(' Ошибка! Коммент не должен превышать длину в 15 символов')
                                continue

                            df_ord.iloc[[i], [1]] = comment

                            sub_row3 = df_ord.loc[df_ord.index == i]

                            q, w, e, r, t, y = sub_row3.iloc[0]['имя'], \
                                               sub_row3.iloc[0]['коммент'], \
                                               sub_row3.iloc[0]['телефон'], \
                                               sub_row3.iloc[0]['дата'], \
                                               sub_row3.iloc[0]['залог'], \
                                               sub_row3.iloc[0]['описание заказа']

                            message = f'#заказ\n' \
                                      f'Имя: #{q}\n' \
                                      f'Инф: {w}\n' \
                                      f'Тел: +7{e}\n' \
                                      f'Дата: {r}\n' \
                                      f'Залог: {t} руб.\n' \
                                      f'--> ( {y} )'

                            token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                            chatID = '-1001820758497'
                            bot = telebot.TeleBot(token)
                            bot.edit_message_text(text=message, chat_id=chatID, message_id=sub_row3.iloc[0]['msg_id'])

                            print('\n' + '\t' + '<<<<<< Коммент изменен! >>>>>>' + '\n')

                        elif choice_ed == 3:
                            while True:
                                try:
                                    deposit = int(input('\n' + ' ВВЕДИТЕ новый залог: '))
                                    if deposit == 0:
                                        break
                                    if str(deposit)[:1] == '0':
                                        raise ValueError(' Залог не должен начинаться с нуля')
                                    deposit = int(deposit)
                                    df_ord.iloc[[i], [2]] = deposit

                                    sub_row = df_ord.loc[df_ord.index == i]

                                    q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                       sub_row.iloc[0]['коммент'], \
                                                       sub_row.iloc[0]['телефон'], \
                                                       sub_row.iloc[0]['дата'], \
                                                       sub_row.iloc[0]['залог'], \
                                                       sub_row.iloc[0]['описание заказа']

                                    message = f'#заказ\n' \
                                              f'Имя: #{q}\n' \
                                              f'Инф: {w}\n' \
                                              f'Тел: +7{e}\n' \
                                              f'Дата: {r}\n' \
                                              f'Залог: {t} руб.\n' \
                                              f'--> ( {y} )'

                                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                    chatID = '-1001820758497'
                                    bot = telebot.TeleBot(token)
                                    bot.edit_message_text(text=message, chat_id=chatID,
                                                          message_id=sub_row.iloc[0]['msg_id'])

                                    print('\n' + '\t' + '<<<<<< Залог изменен! >>>>>>' + '\n')

                                except ValueError:
                                    deposit = str('0')
                                    df_ord.iloc[[i], [2]] = deposit

                                    sub_row = df_ord.loc[df_ord.index == i]

                                    q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                       sub_row.iloc[0]['коммент'], \
                                                       sub_row.iloc[0]['телефон'], \
                                                       sub_row.iloc[0]['дата'], \
                                                       sub_row.iloc[0]['залог'], \
                                                       sub_row.iloc[0]['описание заказа']

                                    message = f'#заказ\n' \
                                              f'Имя: #{q}\n' \
                                              f'Инф: {w}\n' \
                                              f'Тел: +7{e}\n' \
                                              f'Дата: {r}\n' \
                                              f'Залог: {t} руб.\n' \
                                              f'--> ( {y} )'

                                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                    chatID = '-1001820758497'
                                    bot = telebot.TeleBot(token)
                                    bot.edit_message_text(text=message, chat_id=chatID,
                                                          message_id=sub_row.iloc[0]['msg_id'])

                                    print('\n' + '\t' + '<<<<<< Залог изменен! >>>>>>' + '\n')
                                    break
                                else:
                                    break

                            if deposit == 0:
                                continue

                        elif choice_ed == 4:
                            while True:
                                phone = str(input('\n' + ' ВВЕДИТЕ новый телефон: ')) or '9280000000'
                                if phone == '0' or phone == 'выйти' or phone == 'назад' or phone == '9280000000':
                                    break
                                elif phone.isdigit() is True and len(phone) == 10:

                                    df_ord.iloc[[i], [4]] = phone

                                    sub_row = df_ord.loc[df_ord.index == i]

                                    q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                       sub_row.iloc[0]['коммент'], \
                                                       sub_row.iloc[0]['телефон'], \
                                                       sub_row.iloc[0]['дата'], \
                                                       sub_row.iloc[0]['залог'], \
                                                       sub_row.iloc[0]['описание заказа']

                                    message = f'#заказ\n' \
                                              f'Имя: #{q}\n' \
                                              f'Инф: {w}\n' \
                                              f'Тел: +7{e}\n' \
                                              f'Дата: {r}\n' \
                                              f'Залог: {t} руб.\n' \
                                              f'--> ( {y} )'

                                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                    chatID = '-1001820758497'
                                    bot = telebot.TeleBot(token)
                                    bot.edit_message_text(text=message, chat_id=chatID,
                                                          message_id=sub_row.iloc[0]['msg_id'])

                                    print('\n' + '\t' + '<<<<<< Телефон изменен! >>>>>>' + '\n')
                                    break
                                else:
                                    print(' Ошибка! Телефон должен состоять из 10 цифр')
                                    continue

                            if phone == '0' or phone == 'выйти' or phone == 'назад':
                                continue

                        elif choice_ed == 5:
                            while True:
                                descr = input('\n' + ' ВВЕДИТЕ новое описание заказа: ').lower()
                                if descr == '0' or descr == 'выйти' or descr == 'назад':
                                    break
                                if len(descr) > 75:
                                    print(' Ошибка! Описание не должно превышать длину в 15 символов')
                                    continue

                                df_ord.iloc[[i], [5]] = descr

                                sub_row = df_ord.loc[df_ord.index == i]

                                q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                   sub_row.iloc[0]['коммент'], \
                                                   sub_row.iloc[0]['телефон'], \
                                                   sub_row.iloc[0]['дата'], \
                                                   sub_row.iloc[0]['залог'], \
                                                   sub_row.iloc[0]['описание заказа']

                                message = f'#заказ\n' \
                                          f'Имя: #{q}\n' \
                                          f'Инф: {w}\n' \
                                          f'Тел: +7{e}\n' \
                                          f'Дата: {r}\n' \
                                          f'Залог: {t} руб.\n' \
                                          f'--> ( {y} )'

                                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                chatID = '-1001820758497'
                                bot = telebot.TeleBot(token)
                                bot.edit_message_text(text=message, chat_id=chatID,
                                                      message_id=sub_row.iloc[0]['msg_id'])

                                print('\n' + '\t' + '<<<<<< Описание заказа изменено! >>>>>>' + '\n')
                                break

                            if descr == '0' or descr == 'выйти' or descr == 'назад':
                                continue

                        df_ord.to_csv(path_ord_csv, index=False)
                        df_ord.to_excel(path_ord_excel, sheet_name="ЗАКАЗЫ ОТ КЛИЕНТОВ")
                        format_excel(doc_orders, sheet_orders)
                        adjust_excel(doc_orders, sheet_orders)
                        break

                elif is_pressed('g'):
                    while True:
                        search = input('\n' + ' ВВЕДИТЕ имя | коммент | телефон | дата: ')
                        search1 = search[:1].upper() + search[1:].lower()
                        search2 = search.lower()

                        if search == '0' or search == 'выйти' or search == 'назад':
                            print('\n')
                            break

                        else:
                            pd.set_option('display.max_rows', None)
                            pd.set_option('display.max_columns', None)
                            df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
                            df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']
                            df_ord_copy = df_ord.copy()
                            df_ord_copy.drop('msg_id', axis=1, inplace=True)

                            res_df = df_ord_copy.loc[
                                (df_ord_copy.имя == search) | (df_ord_copy.имя == search1) | (
                                        df_ord_copy.имя == search2)
                                | (df_ord_copy.коммент == search) | (df_ord_copy.коммент == search2)
                                | (df_ord_copy.дата == search)
                                | (df_ord_copy.телефон.apply(lambda x: str(x)) == search)]
                            if res_df.empty:
                                print(' Не найдено! Попробуйте снова\n')
                                continue
                            else:
                                print('\n')
                                print(res_df)

                elif is_pressed('l'):
                    while True:
                        name = input('\n' + ' ВВЕДИТЕ имя: ')
                        if name == '0' or name == 'выйти' or name == 'назад':
                            print('\n')
                            break
                        if len(name) < 2 or len(name) > 15:
                            print(' Ошибка! Длина имени должна быть от 2 до 15 символов')
                        name = name[:1].upper() + name[1:].lower()
                        if not name.isalpha():
                            print(' Ошибка! Имя не должно содержать цифры и спец. символы')
                            continue
                        break

                    if name == '0' or name == 'выйти' or name == 'назад':
                        continue

                    comment = input(' ВВЕДИТЕ коммент: ').lower() or '- - -'

                    if comment == '0' or comment == 'выйти' or comment == 'назад':
                        print('\n')
                        continue
                    if len(comment) > 15:
                        print(' Ошибка! Коммент не должен превышать длину в 15 символов')
                        continue

                    while True:
                        descr = input(' ВВЕДИТЕ описание заказа: ').lower()
                        if descr == '' or descr == ' ':
                            continue
                        if descr == '0' or descr == 'выйти' or descr == 'назад':
                            print('\n')
                            break
                        if len(descr) > 75:
                            print(' Ошибка! Описание не должно превышать длину в 75 символов')
                            continue
                        break

                    if descr == '0' or descr == 'выйти' or descr == 'назад':
                        continue

                    while True:
                        try:
                            deposit = int(input(' ВВЕДИТЕ залог: '))
                            if deposit == 0:
                                print('\n')
                                break
                            if str(deposit)[:1] == '0':
                                raise ValueError(' Залог не должен начинаться с нуля')
                            deposit = int(deposit)
                        except ValueError:
                            deposit = str('0')
                            break
                        else:
                            break

                    if deposit == 0:
                        continue

                    while True:
                        phone = str(input(' ВВЕДИТЕ телефон: ')) or '- - -'

                        lnm = list(df_friends['имя'])
                        lph = list(df_friends['телефон'])

                        if name in lnm and not phone.isdigit():
                            i = lnm.index(name)
                            phone = str(lph[i])
                        if phone == '0' or phone == 'выйти' or phone == 'назад' or phone == '- - -':
                            print('\n')
                            break
                        elif phone.isdigit() is True and len(phone) == 10:
                            break
                        else:
                            print(' Ошибка! Телефон должен состоять из 10 цифр')
                            continue

                    if phone == '0' or phone == 'выйти' or phone == 'назад':
                        continue

                    message = f'#заказ\n' \
                              f'Имя: {name}\n' \
                              f'Инф: {comment}\n' \
                              f'Тел: +7{phone}\n' \
                              f'Дата: {day}\n' \
                              f'Залог: {deposit} руб.\n' \
                              f'--> ( {descr} )'
                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                    chatID = '-1001820758497'
                    bot = telebot.TeleBot(token)
                    m = bot.send_message(chat_id=chatID, text=message)
                    msg_id = m.message_id

                    '''print('\n')
                    print(dataframe.loc[dataframe.index == inx])'''
                    '''sub_row = df_ord.loc[df_ord.index == name]'''

                    fcv = open(path_ord_csv, 'a', encoding=enc)
                    fcv.write(
                        '{0},{1},{2},{3},{4},{5},{6}\n'.format(name,
                                                               comment,
                                                               deposit,
                                                               day,
                                                               phone,
                                                               descr,
                                                               msg_id))
                    fcv.close()

                    '''message = f'#заказ\n' \
                              f'Имя: {name}\nИнф: {comment}\nТел: +7{phone}\nДата: {day}\n' \
                              f'Залог: {deposit} руб.\n--> ( {descr} )'
                    apiToken = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                    chatID = '-1001820758497'
                    apiURL = f'https://api.telegram.org/bot{apiToken}/sendMessage'''

                    '''https://api.telegram.org/bot6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4/getUpdates'''
                    '''-820748740'''

                    '''try:
                        response = requests.post(apiURL, json={'chat_id': chatID, 'text': message})
                        print(response.text)
                    except Exception as e:
                        print(e)'''

                    pd.set_option('display.max_rows', None)
                    pd.set_option('display.max_columns', None)
                    df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
                    df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']

                    df_ord.to_csv(path_ord_csv, index=False)
                    df_ord.to_excel(path_ord_excel, sheet_name="ЗАКАЗЫ ОТ КЛИЕНТОВ")
                    format_excel(doc_orders, sheet_orders)
                    adjust_excel(doc_orders, sheet_orders)

                    print('\n')
                    print(df_ord.iloc[-1])
                    print('\n')

                elif is_pressed('n'):
                    pd.set_option('display.max_rows', None)
                    pd.set_option('display.max_columns', None)
                    df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
                    df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']
                    df_ord_copy = df_ord.copy()
                    df_ord_copy.drop('msg_id', axis=1, inplace=True)
                    print('\n')
                    print(df_ord_copy)
                    print('\n')
                    if keyboard.wait('q'):
                        print(' <--')
                        break

                elif is_pressed('e'):
                    while True:
                        pd.set_option('display.max_rows', None)
                        pd.set_option('display.max_columns', None)
                        df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
                        df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']
                        df_ord_copy = df_ord.copy()
                        df_ord_copy.drop('msg_id', axis=1, inplace=True)

                        search = input('\n' + ' ВВЕДИТЕ имя | коммент | телефон | дата | "все" | "инд": ')
                        search1 = search[:1].upper() + search[1:].lower()
                        search2 = search.lower()

                        if search == '0' or search == 'выйти' or search == 'назад':
                            print('\n')
                            break

                        elif search == 'все':
                            print('\n')
                            print(df_ord_copy)
                            continue

                        elif search == 'инд':
                            ind = int(input(' ВВЕДИТЕ индекс: '))

                            if ind not in df_ord.index:
                                raise IndexError

                            df_ord.drop(df_ord[df_ord.index == ind].index, inplace=True)
                            print('\n' + '\t' + '<<<<<< Удалено! >>>>>>')
                            df_ord.to_csv(path_ord_csv, index=False)
                            df_ord.to_excel(path_ord_excel, sheet_name="ЗАКАЗЫ ОТ КЛИЕНТОВ")
                            continue

                        else:
                            res_df = df_ord_copy.loc[(df_ord_copy.имя == search) | (df_ord_copy.имя == search1)
                                                     | (df_ord_copy.имя == search2)
                                                     | (df_ord_copy.коммент == search2)
                                                     | (df_ord_copy.дата == search)
                                                     | (df_ord_copy.телефон.apply(lambda x: str(x)) == search)]

                            if res_df.empty:
                                print(' Не найдено! Попробуйте снова\n')
                                continue

                            else:
                                print('\n')
                                print(res_df)
                                continue

        except ValueError:
            print('\n' + ' Ошибка! Введите цифровое значение' + '\n')
        except IndexError:
            print('\n' + ' Ошибка! Такого индекса не существует' + '\n')
        except PermissionError:
            print('\n' + ' Ошибка! Закройте excel файл' + '\n')


def sort_func():
    while True:
        print(' НАЖМИТЕ "1" имя | "2" коммент | "4" долг')
        while True:
            if is_pressed('q'):
                break

            if is_pressed('1'):
                df8 = df().sort_values('имя', ascending=True)
                print('\n')
                print(df8)
                print('\n')
                break

            elif is_pressed('2'):
                df8 = df().sort_values('коммент', ascending=True)
                print('\n')
                print(df8)
                print('\n')
                break

            elif is_pressed('4'):
                df8 = df().sort_values('долг', ascending=True)
                print('\n')
                print(df8)
                print('\n')
                break

        if is_pressed('q'):
            break
        if keyboard.wait('p'):
            print(' <--')
            continue


def edit_func(debt=None):
    while True:
        try:
            choice_edit = int(input('\n'
                                    '        \'1\' ДАННЫЕ (о долгах)\n'
                                    '        \'2\' ПАРОЛЬ\n'
                                    '        \'3\' ДОБАВИТЬ ДРУГА\n'
                                    '        \'4\' УДАЛИТЬ ДРУГА\n'
                                    '        \'5\' ТАБЛИЦА ДРУ3ЕЙ\n'
                                    '    --> '))

            if choice_edit == 0:
                break

            elif choice_edit == 1:
                try:
                    while True:
                        pd.set_option('display.max_rows', None)
                        pd.set_option('display.max_columns', None)
                        dataframe2 = pd.read_csv(path_csv, engine='python', encoding=enc)
                        dataframe2.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
                        dataframe2_copy = dataframe2.copy()
                        dataframe2_copy.drop('msg_id', axis=1, inplace=True)

                        print('\n')
                        print(dataframe2_copy.tail(15))
                        print('\n')

                        i = int(input(' ВВЕДИТЕ индекс: '))

                        if i == 0:
                            break

                        '''ed_row = df_ord.loc[df_ord.index == i]'''

                        choice_ed = int(input('\n'
                                              '        \'1\' ИМЯ\n'
                                              '        \'2\' КОММЕНТ\n'
                                              '        \'3\' ДОЛГ\n'
                                              '        \'4\' ТЕЛЕФОН\n'
                                              '        \'5\' ОПИСАНИЕ\n'
                                              '    --> '))

                        if choice_ed == 1:
                            while True:
                                name = input('\n' + ' ВВЕДИТЕ новое имя: ')
                                if name == '0' or name == 'выйти' or name == 'назад':
                                    break
                                name = name[:1].upper() + name[1:].lower()
                                if not name.isalpha():
                                    print(' Ошибка! Имя не должно содержать цифры и спец. символы')
                                    continue
                                dataframe2.iloc[[i], [0]] = name

                                sub_row = dataframe2.loc[dataframe2.index == i]

                                q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                   sub_row.iloc[0]['коммент'], \
                                                   sub_row.iloc[0]['телефон'], \
                                                   sub_row.iloc[0]['дата'], \
                                                   sub_row.iloc[0]['описание'], \
                                                   sub_row.iloc[0]['долг']

                                message = f'#долг\n' \
                                          f'Имя: #{q}\n' \
                                          f'Инф: {w}\n' \
                                          f'Тел: +7{e}\n' \
                                          f'Дата: {r}\n' \
                                          f': {t}\n' \
                                          f'--> {y} руб.'

                                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                chatID = '-1001820758497'
                                bot = telebot.TeleBot(token)
                                bot.edit_message_text(text=message, chat_id=chatID,
                                                      message_id=sub_row.iloc[0]['msg_id'])

                                print('\n' + '\t' + '<<<<<< Имя изменено! >>>>>>')
                                break

                            if name == '0' or name == 'выйти' or name == 'назад':
                                continue

                        elif choice_ed == 2:

                            comment = input('\n' + ' ВВЕДИТЕ новый коммент: ').lower() or '- - -'

                            if comment == '0' or comment == 'выйти' or comment == 'назад':
                                continue
                            if len(comment) > 15:
                                print(' Ошибка! Коммент не должен превышать длину в 15 символов')
                                continue
                            dataframe2.iloc[[i], [1]] = comment

                            sub_row2 = dataframe2.loc[dataframe2.index == i]

                            q, w, e, r, t, y = sub_row2.iloc[0]['имя'], \
                                               sub_row2.iloc[0]['коммент'], \
                                               sub_row2.iloc[0]['телефон'], \
                                               sub_row2.iloc[0]['дата'], \
                                               sub_row2.iloc[0]['описание'], \
                                               sub_row2.iloc[0]['долг']

                            message = f'#долг\n' \
                                      f'Имя: #{q}\n' \
                                      f'Инф: {w}\n' \
                                      f'Тел: +7{e}\n' \
                                      f'Дата: {r}\n' \
                                      f': {t}\n' \
                                      f'--> {y} руб.'

                            token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                            chatID = '-1001820758497'
                            bot = telebot.TeleBot(token)
                            bot.edit_message_text(text=message, chat_id=chatID, message_id=sub_row2.iloc[0]['msg_id'])

                            print('\n' + '\t' + '<<<<<< Коммент изменен! >>>>>>')

                        elif choice_ed == 3:
                            while True:
                                try:
                                    debt = int(input('\n' + ' ВВЕДИТЕ новый долг: '))
                                    if debt == 0:
                                        break
                                    if str(debt)[:1] == '0':
                                        raise ValueError(' Долг не должен начинаться с нуля')
                                except ValueError:
                                    print(' Ошибка! Долг должен состоять из цифр')
                                    continue
                                else:
                                    dataframe2.iloc[[i], [2]] = debt

                                    sub_row = dataframe2.loc[dataframe2.index == i]

                                    q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                       sub_row.iloc[0]['коммент'], \
                                                       sub_row.iloc[0]['телефон'], \
                                                       sub_row.iloc[0]['дата'], \
                                                       sub_row.iloc[0]['описание'], \
                                                       sub_row.iloc[0]['долг']

                                    message = f'#долг\n' \
                                              f'Имя: #{q}\n' \
                                              f'Инф: {w}\n' \
                                              f'Тел: +7{e}\n' \
                                              f'Дата: {r}\n' \
                                              f': {t}\n' \
                                              f'--> {y} руб.'

                                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                    chatID = '-1001820758497'
                                    bot = telebot.TeleBot(token)
                                    bot.edit_message_text(text=message, chat_id=chatID,
                                                          message_id=sub_row.iloc[0]['msg_id'])

                                    print('\n' + '\t' + '<<<<<< Долг изменен! >>>>>>' + '\n')
                                    break

                            if debt == 0:
                                continue

                        elif choice_ed == 4:
                            while True:
                                phone = str(input('\n' + ' ВВЕДИТЕ новый телефон: ')) or '9280000000'
                                if phone == '0' or phone == 'выйти' or phone == 'назад' or phone == '9280000000':
                                    break
                                elif phone.isdigit() is True and len(phone) == 10:

                                    dataframe2.iloc[[i], [4]] = phone

                                    sub_row = dataframe2.loc[dataframe2.index == i]

                                    q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                       sub_row.iloc[0]['коммент'], \
                                                       sub_row.iloc[0]['телефон'], \
                                                       sub_row.iloc[0]['дата'], \
                                                       sub_row.iloc[0]['описание'], \
                                                       sub_row.iloc[0]['долг']

                                    message = f'#долг\n' \
                                              f'Имя: #{q}\n' \
                                              f'Инф: {w}\n' \
                                              f'Тел: +7{e}\n' \
                                              f'Дата: {r}\n' \
                                              f': {t}\n' \
                                              f'--> {y} руб.'

                                    token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                    chatID = '-1001820758497'
                                    bot = telebot.TeleBot(token)
                                    bot.edit_message_text(text=message, chat_id=chatID,
                                                          message_id=sub_row.iloc[0]['msg_id'])

                                    print('\n' + '\t' + '<<<<<< Телефон изменен! >>>>>>')
                                    break
                                else:
                                    print(' Ошибка! Телефон должен состоять из 10 цифр')
                                    continue

                            if phone == '0' or phone == 'выйти' or phone == 'назад':
                                continue

                        elif choice_ed == 5:
                            while True:
                                descr = input('\n' + ' ВВЕДИТЕ новое описание: ').lower()
                                if descr == '0' or descr == 'выйти' or descr == 'назад':
                                    break
                                if len(descr) > 75:
                                    print(' Ошибка! Описание не должно превышать длину в 15 символов')
                                    continue

                                dataframe2.iloc[[i], [5]] = descr

                                sub_row = dataframe2.loc[dataframe2.index == i]

                                q, w, e, r, t, y = sub_row.iloc[0]['имя'], \
                                                   sub_row.iloc[0]['коммент'], \
                                                   sub_row.iloc[0]['телефон'], \
                                                   sub_row.iloc[0]['дата'], \
                                                   sub_row.iloc[0]['описание'], \
                                                   sub_row.iloc[0]['долг']

                                message = f'#долг\n' \
                                          f'Имя: #{q}\n' \
                                          f'Инф: {w}\n' \
                                          f'Тел: +7{e}\n' \
                                          f'Дата: {r}\n' \
                                          f': {t}\n' \
                                          f'--> {y} руб.'

                                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                                chatID = '-1001820758497'
                                bot = telebot.TeleBot(token)
                                bot.edit_message_text(text=message, chat_id=chatID,
                                                      message_id=sub_row.iloc[0]['msg_id'])

                                print('\n' + '\t' + '<<<<<< Описание изменено! >>>>>>' + '\n')
                                break

                            if descr == '0' or descr == 'выйти' or descr == 'назад':
                                continue

                        dataframe2.to_csv(path_csv, index=False)
                        dataframe2.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")
                        format_excel(doc_loans, sheet_loans)
                        adjust_excel(doc_loans, sheet_loans)
                        break

                except ValueError:
                    print('\n' + ' Ошибка! Введите цифровое значение')
                except IndexError:
                    print('\n' + ' Ошибка! Такого индекса не существует')
                except PermissionError:
                    print('\n' + ' Ошибка! Закройте excel файл')

            elif choice_edit == 2:
                while True:
                    password = raw_input('\n' + ' ВВЕДИТЕ новый пароль: ')

                    if password == '0':
                        print(' Выход...')
                        break

                    if len(password) < 8:
                        print(" Пароль слишком короткий")

                    elif re.search('\\d', password) is None:
                        print(" Пароль должен содержать минимум одну цифру")

                    elif re.search('[А-Я]', password) is None:
                        print(" Пароль должен содержать минимум одну заглавную букву")

                    elif ' ' in password:
                        print(' Пароль содержит пробелы')

                    elif password == '0':
                        print(' Выход...')

                    else:
                        print('\n' + '\t' + '<<<<<< Пароль изменен! >>>>>>')

                        with open(path_passwd, 'w') as f:
                            enc_passwd = cryptocode.encrypt(password, 'wow')
                            f.write(enc_passwd)

                        break

            elif choice_edit == 3:
                while True:
                    pd.set_option('display.max_rows', None)
                    pd.set_option('display.max_columns', None)
                    df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
                    df_friends.columns = ['имя', 'телефон']
                    lnm = list(df_friends.имя)
                    lph = list(df_friends.телефон)
                    try:
                        while True:
                            print('\n')
                            name = input(' ВВЕДИТЕ имя: ')
                            if name == '0' or name == 'выйти' or name == 'назад':
                                break
                            name = name[:1].upper() + name[1:].lower()
                            if name in lnm:
                                print(' Ошибка! Такое имя уже существует')
                                continue
                            if not name.isalpha():
                                print(' Ошибка! Имя не должно содержать цифры и спец. символы')
                                continue
                            break

                        if name == '0' or name == 'выйти' or name == 'назад':
                            break

                        while True:
                            lph1 = list(map(str, lph))
                            phone = str(input(' ВВЕДИТЕ телефон: '))

                            if phone == '0' or phone == 'выйти' or phone == 'назад':
                                break
                            if phone in lph1:
                                print(' Ошибка! Такой телефон уже существует')
                            elif phone.isdigit() is True and len(phone) == 10:
                                break
                            else:
                                print(' Ошибка! Телефон должен состоять из 10 цифр')
                                continue

                        if phone == '0' or phone == 'выйти' or phone == 'назад':
                            continue

                        print('\n' + '\t' + '<<<<<< Добавлено! >>>>>>')

                        fcv = open(path_friends_csv, 'a', encoding=enc)
                        fcv.write('{0},{1}\n'.format(name, phone))
                        fcv.close()
                    except ValueError:
                        print('\n' + ' Ошибка! Неверное значение!')

            elif choice_edit == 4:
                while True:
                    try:
                        pd.set_option('display.max_rows', None)
                        pd.set_option('display.max_columns', None)
                        df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
                        df_friends.columns = ['имя', 'телефон']
                        print(' ')
                        print(df_friends.tail(10))
                        print(' ')

                        search = input(' ВВЕДИТЕ имя | "инд": ')
                        search1 = search[:1].upper() + search[1:].lower()
                        search2 = search.lower()

                        if search == '0' or search == 'выйти' or search == 'назад':
                            break

                        elif search == 'инд':
                            ind = int(input(' ВВЕДИТЕ индекс: '))

                            if ind not in df_friends.index:
                                raise IndexError

                            df_friends.drop(df_friends[df_friends.index == ind].index, inplace=True)
                            print('\n' + '\t' + '<<<<<< Удалено! >>>>>>')
                            df_friends.to_csv(path_friends_csv, index=False)
                            continue

                        else:
                            pd.set_option('display.max_rows', None)
                            df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
                            df_friends.columns = ['имя', 'телефон']

                            res_dt = df_friends.loc[(df_friends.имя == search) | (df_friends.имя == search1)
                                                    | (df_friends.имя == search2)
                                                    | (df_friends.телефон.apply(lambda x: str(x)) == search)]
                            if res_dt.empty:
                                print(' Не найдено! Попробуйте снова\n')
                                continue
                            else:
                                print('\n')
                                print(res_dt)
                                print('\n')

                                ch = int(input(' УДАЛИТЬ ЭТИ ДАННЫЕ?\n'
                                               ' \'1\' ДА\n'
                                               ' \'0\' ОТМЕНА\n'
                                               ' --> '))
                                if ch == 1:
                                    df_friends.drop(df_friends[(df_friends.имя == search) | (df_friends.имя == search1)
                                                               | (df_friends.имя == search2)].index, inplace=True)

                                    print('\n' + '\t' + '<<<<<< Удалено! >>>>>>')
                                    df_friends.to_csv(path_friends_csv, index=False)
                                elif ch == 0:
                                    continue
                    except ValueError:
                        print(' Ошибка! Индекс должен состоять из цифр\n')
                        continue
                    except IndexError:
                        print(' Ошибка! Такого индекса не существует\n')
                        continue
                    except PermissionError:
                        print('\n' + ' Ошибка! Закройте файл')
                        continue

            elif choice_edit == 5:
                pd.set_option('display.max_rows', None)
                pd.set_option('display.max_columns', None)
                df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
                df_friends.columns = ['имя', 'телефон']
                print('\n')
                print(df_friends)

        except ValueError:
            print(' Ошибка! Неверное значение!')


def bar_func():
    while True:
        print(' НАЖМИТЕ "1" имя | "2" коммент | "3" тел ')

        if is_pressed('q'):
            break

        while True:
            if is_pressed('q'):
                break
            if is_pressed('1'):
                df3 = df()
                df3 = df3.replace(np.nan, 0)
                sums = df3.groupby('имя')['долг'].sum()

                # Figure Size
                fig, ax = plt1.subplots(figsize=(14, 7.5))

                names = []
                debts = []
                for key, val in sums.items():
                    names.append(key)
                    debts.append(val)

                # Horizontal Bar Plot
                ax.barh(names, debts)

                # Remove axes splines
                for s in ['top', 'bottom', 'left', 'right']:
                    ax.spines[s].set_visible(False)

                # Remove x, y Ticks
                ax.xaxis.set_ticks_position('none')
                ax.yaxis.set_ticks_position('none')

                # Add padding between axes and labels
                ax.xaxis.set_tick_params(pad=5)
                ax.yaxis.set_tick_params(pad=10)

                # Add x, y gridlines
                ax.grid(b=True, color='grey',
                        linestyle='-.', linewidth=0.5,
                        alpha=0.2)

                # Show top values
                ax.invert_yaxis()

                # Add annotation to bars
                for i in ax.patches:
                    plt1.text(i.get_width() + 0.2, i.get_y() + 0.5,
                              str(round((i.get_width()), 2)),
                              fontsize=7.5, fontweight='bold',
                              color='red')

                # Add Plot Title
                ax.set_title('Сумма долга по каждому имени',
                             loc='left', )

                # Add Text watermark
                fig.text(0.9, 0.15, 'yahyagadzhi@gmail.com', fontsize=12,
                         color='grey', ha='right', va='bottom',
                         alpha=0.7)

                # Show Plot
                plt1.show()
                break

            elif is_pressed('2'):
                df4 = df()
                df4 = df4.replace(np.nan, 0)
                df4['имя_коммент'] = df4.имя + ' (' + df4.коммент + ')'
                sums = df4.groupby('имя_коммент')['долг'].sum()

                # Figure Size
                fig, ax = plt1.subplots(figsize=(14, 7.5))

                names = []
                debts = []
                for key, val in sums.items():
                    names.append(key)
                    debts.append(val)

                # Horizontal Bar Plot
                ax.barh(names, debts)

                # Remove axes splines
                for s in ['top', 'bottom', 'left', 'right']:
                    ax.spines[s].set_visible(False)

                # Remove x, y Ticks
                ax.xaxis.set_ticks_position('none')
                ax.yaxis.set_ticks_position('none')

                # Add padding between axes and labels
                ax.xaxis.set_tick_params(pad=5)
                ax.yaxis.set_tick_params(pad=10)

                # Add x, y gridlines
                ax.grid(b=True, color='grey',
                        linestyle='-.', linewidth=0.5,
                        alpha=0.2)

                # Show top values
                ax.invert_yaxis()

                # Add annotation to bars
                for i in ax.patches:
                    plt1.text(i.get_width() + 0.2, i.get_y() + 0.5,
                              str(round((i.get_width()), 2)),
                              fontsize=7.5, fontweight='bold',
                              color='red')

                # Add Plot Title
                ax.set_title('Сумма долга по каждому имя & коммент',
                             loc='left', )

                # Add Text watermark
                fig.text(0.9, 0.15, 'yahyagadzhi@gmail.com', fontsize=12,
                         color='grey', ha='right', va='bottom',
                         alpha=0.7)

                # Show Plot
                plt1.show()
                break

            elif is_pressed('3'):
                df5 = df()
                df5 = df5.replace(np.nan, 0)
                df5['имя_коммент_телефон'] = df5.имя + ' (' + df5.коммент + ') ' + '+7' + df5.телефон.apply(
                    lambda x: str(x))
                sums = df5.groupby('имя_коммент_телефон')['долг'].sum()

                # Figure Size
                fig, ax = plt1.subplots(figsize=(14, 7.5))

                names = []
                debts = []
                for key, val in sums.items():
                    names.append(key)
                    debts.append(val)

                # Horizontal Bar Plot
                ax.barh(names, debts)

                # Remove axes splines
                for s in ['top', 'bottom', 'left', 'right']:
                    ax.spines[s].set_visible(False)

                # Remove x, y Ticks
                ax.xaxis.set_ticks_position('none')
                ax.yaxis.set_ticks_position('none')

                # Add padding between axes and labels
                ax.xaxis.set_tick_params(pad=5)
                ax.yaxis.set_tick_params(pad=10)

                # Add x, y gridlines
                ax.grid(b=True, color='grey',
                        linestyle='-.', linewidth=0.5,
                        alpha=0.2)

                # Show top values
                ax.invert_yaxis()

                # Add annotation to bars
                for i in ax.patches:
                    plt1.text(i.get_width() + 0.2, i.get_y() + 0.5,
                              str(round((i.get_width()), 2)),
                              fontsize=7.5, fontweight='bold',
                              color='red')

                # Add Plot Title
                ax.set_title('Сумма долга по каждому имя & коммент & телефон',
                             loc='left', )

                # Add Text watermark
                fig.text(0.9, 0.15, 'yahyagadzhi@gmail.com', fontsize=12,
                         color='grey', ha='right', va='bottom',
                         alpha=0.7)

                # Show Plot
                plt1.show()
                break


def df():
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    dataframe = pd.read_csv(path_csv, engine='python', encoding=enc)
    dataframe.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
    dataframe['телефон'].apply(lambda x: str(x))
    df_copy = dataframe.copy()
    df_copy.drop('msg_id', axis=1, inplace=True)
    return df_copy


def del_func():
    while True:
        try:
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            dataframe = pd.read_csv(path_csv, engine='python', encoding=enc)
            dataframe.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
            dataframe_copy = dataframe.copy()
            dataframe_copy.drop('msg_id', axis=1, inplace=True)

            print('\n')
            print(dataframe_copy.tail(10))
            print('\n')

            search = input(' ВВЕДИТЕ имя | коммент | телефон | дата | "инд": ')
            search1 = search[:1].upper() + search[1:].lower()
            search2 = search.lower()

            if search == '0' or search == 'выйти' or search == 'назад':
                break

            elif search == 'инд':
                ind = int(input(' ВВЕДИТЕ индекс: '))

                if ind not in dataframe.index:
                    raise IndexError

                sub_row1 = dataframe.loc[dataframe.index == ind]
                k = sub_row1.iloc[0]['долг']

                dataframe.iloc[[ind], [2]] -= k

                sub_row = dataframe.loc[dataframe.index == ind]

                a, b, j = sub_row.iloc[0]['имя'], sub_row.iloc[0]['коммент'], sub_row.iloc[0]['описание']
                c1, d1, p1 = sub_row.iloc[0]['долг'], sub_row.iloc[0]['дата'], sub_row.iloc[0]['телефон']

                dataframe.drop(dataframe[dataframe.index == ind].index, inplace=True)

                day = datetime.today().strftime('%d.%m.%Y')

                message = f'#уддолг\n' \
                          f'Имя: #{a}\nИнф: {b}\nТел: +7{p1}\nДата: {d1}\n: {j}\n--> {k} руб.\n' \
                          f'Дата удал: {day}\n' \
                          f'Остаток: {c1} руб.\n'
                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                chatID = '-1001820758497'
                bot = telebot.TeleBot(token)
                bot.edit_message_text(text=message, chat_id=chatID, message_id=sub_row.iloc[0]['msg_id'])

                print('\n' + '\t' + '<<<<<< Удалено! >>>>>>' + '\n')

                dataframe.to_csv(path_csv, index=False)
                dataframe.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")
                format_excel(doc_loans, sheet_loans)
                adjust_excel(doc_loans, sheet_loans)

                '''print('\n')
                print(dataframe.loc[dataframe.index == inx])'''

                dataframe.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")

                '''message = f'#удаление\n' \
                          f'Имя: #{a}\nИнф: {b}\nДолг: {k}\nДата: {d1}\nТел: +7{p1}\n: {j}\nДата удал: {day}\n' \
                          f'Остаток: {c1} руб.\n'
                apiToken = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                chatID = '-1001820758497'
                apiURL = f'https://api.telegram.org/bot{apiToken}/sendMessage'''

                '''https://api.telegram.org/bot6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4/getUpdates'''

                '''try:
                    response = requests.post(apiURL, json={'chat_id': chatID, 'text': message})
                    print(response.text)
                except Exception as e:
                    print(e)'''

                continue

            else:
                pd.set_option('display.max_rows', None)
                dataframe = pd.read_csv(path_csv, engine='python', encoding=enc)
                dataframe.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']

                res_dt = dataframe.loc[(dataframe.имя == search) | (dataframe.имя == search1)
                                       | (dataframe.имя == search2)
                                       | (dataframe.коммент == search) | (dataframe.коммент == search2)
                                       | (dataframe.дата == search)
                                       | (dataframe.телефон.apply(lambda x: str(x)) == search)]
                if res_dt.empty:
                    print(' Не найдено! Попробуйте снова\n')
                    continue
                else:
                    res_dt_copy = res_dt.copy()
                    res_dt_copy.drop('msg_id', axis=1, inplace=True)
                    print('\n')
                    print(res_dt_copy)
                    total = res_dt.долг.sum()
                    print(f' СУММАРНЫЙ ДОЛГ = {total} руб.')
                    print('\n')

                    ch = int(input(' УДАЛИТЬ ЭТИ ДАННЫЕ?\n'
                                   ' \'1\' ДА\n'
                                   ' \'0\' ОТМЕНА\n'
                                   ' --> '))
                    if ch == 1:
                        dataframe.drop(dataframe[(dataframe.имя == search) | (dataframe.имя == search1)
                                                 | (dataframe.имя == search2)
                                                 | (dataframe.коммент == search) | (dataframe.коммент == search2)
                                                 | (dataframe.дата == search)
                                                 | (dataframe.телефон.apply(lambda x: str(x)) == search)].index,
                                       inplace=True)

                        day1 = datetime.today().strftime('%d.%m.%Y')
                        res_dt.reset_index()
                        for i, row in res_dt.iterrows():
                            message = f'#уддолг\n' \
                                      f'Имя: #{row.имя}\nИнф: {row.коммент}\nТел: +7{row.телефон}\n' \
                                      f'Дата: {row.дата}\n: {row.описание}\n--> {row.долг} руб.\n' \
                                      f'Дата удал: {day1}\n' \
                                      f'Остаток: 0 руб.\n'
                            token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                            chatID = '-1001820758497'
                            bot = telebot.TeleBot(token)
                            bot.edit_message_text(text=message, chat_id=chatID, message_id=row['msg_id'])

                        print('\n' + '\t' + '<<<<<< Удалено! >>>>>>')

                        dataframe.to_csv(path_csv, index=False)
                        dataframe.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")
                        format_excel(doc_loans, sheet_loans)
                        adjust_excel(doc_loans, sheet_loans)
                    elif ch == 0:
                        continue

        except ValueError:
            print(' Ошибка! Индекс должен состоять из цифр\n')
            continue
        except IndexError:
            print(' Ошибка! Такого индекса не существует\n')
            continue
        except PermissionError:
            print('\n' + ' Ошибка! Закройте excel файл')
            continue
        except ApiTelegramException:
            print('\n' + ' ApiTelegramExceptionError!')
            continue


def clean_func():
    print('\n'
          ' Вы уверены?\n'
          ' \'1\' ДА\n'
          ' \'й\' ОТМЕНА')
    while True:
        if is_pressed('q'):
            print(' <--')
            break
        if is_pressed('1'):
            dataframe = pd.read_csv(path_csv, engine='python', encoding=enc)
            dataframe.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
            dataframe.drop(dataframe[dataframe.долг <= 0].index, inplace=True)
            dataframe.to_csv(path_csv, index=False)
            dataframe.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")
            format_excel(doc_loans, sheet_loans)
            adjust_excel(doc_loans, sheet_loans)

            df_orc = pd.read_csv(path_orc_csv, engine='python', encoding=enc)
            df_orc.columns = ['товар', 'количество', 'дата']
            df_orc_copy = df_orc.copy()
            df_orc_copy['дата'] = pd.to_datetime(df_orc_copy['дата'], dayfirst=True)

            days = timedelta(days=14)
            days1 = timedelta(days=35)

            today = datetime.strptime(datetime.today().strftime('%d.%m.%Y'), '%d.%m.%Y')
            dd = today - days

            df_orc_copy.drop(df_orc_copy[df_orc_copy['дата'] <= dd].index, inplace=True)
            df_orc_copy.to_csv(path_orc_csv, index=False)
            df_orc_copy.to_excel(path_orc_excel, sheet_name="ПОПОЛНЕНИЕ ТОВАРА")
            adjust_excel(doc_refills, sheet_refills)

            dataframe_subs = pd.read_csv(path_subs_csv, engine='python', encoding=enc)
            dataframe_subs.columns = ['имя', 'коммент', 'долг', 'изменение', 'дата',
                                      'телефон', 'дата изм.', 'время изм.', 'комментарий к изм.', 'описание']
            dataframe_subs_copy = dataframe_subs.copy()
            dataframe_subs_copy['дата'] = pd.to_datetime(dataframe_subs_copy['дата'])
            dd1 = today - days1
            dataframe_subs_copy['дата'] = pd.to_datetime(dataframe_subs_copy['дата'], dayfirst=True)
            dataframe_subs_copy.drop(dataframe_subs_copy[dataframe_subs_copy['дата'] <= dd1].index, inplace=True)
            dataframe_subs_copy.to_csv(path_subs_csv, index=False)

            df_ord = pd.read_csv(path_ord_csv, engine='python', encoding=enc)
            df_ord.columns = ['имя', 'коммент', 'залог', 'дата', 'телефон', 'описание заказа', 'msg_id']
            df_ord_copy = df_ord.copy()
            df_ord_copy['дата'] = pd.to_datetime(df_ord_copy['дата'], dayfirst=True)
            df_ord_copy.drop(df_ord_copy[df_ord_copy['дата'] <= dd].index, inplace=True)
            df_ord_copy.to_csv(path_ord_csv, index=False)
            df_ord_copy.to_excel(path_ord_excel, sheet_name="ЗАКАЗЫ ОТ КЛИЕНТОВ")
            format_excel(doc_orders, sheet_orders)
            adjust_excel(doc_orders, sheet_orders)

            print('\n' + '\t<<<<<< ОЧИЩЕНО! >>>>>>')
            break


def add_func(debt=None):
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    df_friends = pd.read_csv(path_friends_csv, engine='python', encoding=enc)
    df_friends.columns = ['имя', 'телефон']
    while True:
        while True:
            name = input(' ВВЕДИТЕ имя: ')
            if name == '0' or name == 'выйти' or name == 'назад':
                break
            if len(name) < 2 or len(name) > 15:
                print(' Ошибка! Длина имени должна быть от 2 до 15 символов')

            name = name[:1].upper() + name[1:].lower()
            if not name.isalpha():
                print(' Ошибка! Имя не должно содержать цифры и спец. символы')
                continue
            break

        if name == '0' or name == 'выйти' or name == 'назад':
            break

        comment = input(' ВВЕДИТЕ комментарий: ').lower() or '- - -'

        if comment == '0' or comment == 'выйти' or comment == 'назад':
            continue
        if len(comment) > 15:
            print(' Ошибка! Коммент не должен превышать длину в 15 символов')
            continue

        while True:
            try:
                debt = int(input(' ВВЕДИТЕ долг: '))
                if debt == 0:
                    break
                if str(debt)[:1] == '0':
                    raise ValueError(' Долг не должен начинаться с нуля')
            except ValueError:
                print(' Ошибка! Долг должен состоять из цифр')
                continue
            else:
                break

        if debt == 0:
            continue

        while True:
            phone = str(input(' ВВЕДИТЕ телефон: ')) or '- - -'

            lnm = list(df_friends['имя'])
            lph = list(df_friends['телефон'])

            if name in lnm and not phone.isdigit():
                i = lnm.index(name)
                phone = str(lph[i])

            if phone == '0' or phone == 'выйти' or phone == 'назад' or phone == '- - -':
                break
            elif phone.isdigit() is True and len(phone) == 10:
                break
            else:
                print(' Ошибка! Телефон должен состоять из 10 цифр')
                continue

        if phone == '0' or phone == 'выйти' or phone == 'назад':
            continue

        descr = input(' ВВЕДИТЕ описание: ').lower() or '- - -'

        if descr == '0' or descr == 'выйти' or descr == 'назад':
            continue
        if len(descr) > 75:
            print(' Ошибка! Описание не должно превышать длину в 75 символов')
            continue

        day = datetime.today().strftime('%d.%m.%Y')

        message = f'#долг\n' \
                  f'Имя: #{name}\nИнф: {comment}\nТел: +7{phone}\nДата: {day}\n: {descr}\n--> {debt} руб.'
        token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
        chatID = '-1001820758497'
        bot = telebot.TeleBot(token)
        m = bot.send_message(chat_id=chatID, text=message)
        msg_id = m.message_id

        fcv = open(path_csv, 'a', encoding=enc)
        fcv.write('{0},{1},{2},{3},{4},{5},{6}\n'.format(name, comment, debt, day, phone, descr, msg_id))
        fcv.close()

        '''message = f'#долг\n' \
                  f'Имя: #{name}\nИнф: {comment}\nТел: +7{phone}\nДата: {day}\n: {descr}\n--> {debt} руб.'
        apiToken = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
        chatID = '-1001820758497'
        apiURL = f'https://api.telegram.org/bot{apiToken}/sendMessage'''

        '''https://api.telegram.org/bot6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4/getUpdates'''

        '''try:
            response = requests.post(apiURL, json={'chat_id': chatID, 'text': message})
            print(response.text)
        except Exception as e:
            print(e)'''

        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        dataframe1 = pd.read_csv(path_csv, engine='python', encoding=enc)
        dataframe1.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
        dataframe1_copy = dataframe1.copy()
        dataframe1_copy.drop('msg_id', axis=1, inplace=True)

        dataframe1_copy.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")
        format_excel(doc_loans, sheet_loans)
        adjust_excel(doc_loans, sheet_loans)

        print('\n')
        print(dataframe1_copy.tail(5))
        print('\n')

        continue


def sub_func():
    while True:
        try:
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            dataframe = pd.read_csv(path_csv, engine='python', encoding=enc)
            dataframe.columns = ['имя', 'коммент', 'долг', 'дата', 'телефон', 'описание', 'msg_id']
            dataframe_copy = dataframe.copy()
            dataframe_copy.drop('msg_id', axis=1, inplace=True)

            choice = int(input('\n'
                               '        \'1\' СПИСАТЬ ПО ИНДЕКСУ\n'
                               '        \'2\' ПОИСК\n'
                               '        \'3\' ТАБЛИЦА СПИСАНИЙ\n'
                               '    --> '))

            if choice == 0:
                break

            elif choice == 2:
                search_func()

            elif choice == 1:
                print('\n')
                print(dataframe_copy.tail(10))
                print('\n')

                day, time = datetime.today().strftime('%d.%m.%Y'), datetime.now().strftime('%H:%M:%S')

                inx = int(input(' ВВЕДИТЕ индекс: '))

                num = int(input(' ВВЕДИТЕ число: '))

                if num == 0:
                    break

                comm = input(' ВВЕДИТЕ комментарий к изм: ') or '- - -'

                if comm == '0' or comm == 'выйти' or comm == 'назад':
                    break
                if len(comm) > 30:
                    print(' Ошибка! Комментарий не должен превышать длину в 30 символов')
                    continue

                sub_row1 = dataframe.loc[dataframe.index == inx]
                k = sub_row1.iloc[0]['долг']

                dataframe.iloc[[inx], [2]] -= num
                dataframe.to_csv(path_csv, index=False)
                '''print('\n')
                print(dataframe.loc[dataframe.index == inx])'''

                dataframe.to_excel(path_excel, sheet_name="ДОЛГИ КЛИЕНТОВ")

                sub_row = dataframe.loc[dataframe.index == inx]

                a, b, j = sub_row.iloc[0]['имя'], sub_row.iloc[0]['коммент'], sub_row.iloc[0]['описание']
                c1, d1, p1 = sub_row.iloc[0]['долг'], sub_row.iloc[0]['дата'], sub_row.iloc[0]['телефон']

                message = f'#долг\n' \
                          f'Имя: #{a}\nИнф: {b}\nДолг: {k}\nДата: {d1}\nТел: +7{p1}\n: {j}\n' \
                          f'#списание\n' \
                          f'Дата спис: {day}\n' \
                          f'Инф к спис: {comm}\n' \
                          f'--> ( - {num} руб. )\n' \
                          f'Остаток: {c1} руб.\n '
                token = '6176587339:AAEoi7OpgOFgwZXqjSCpZun-uyCzaEhvAq4'
                chatID = '-1001820758497'
                bot = telebot.TeleBot(token)
                bot.edit_message_text(text=message, chat_id=chatID, message_id=int(sub_row['msg_id']))

                fcv = open(path_subs_csv, 'a', encoding=enc)
                fcv.write(
                    '{0},{1},{2},-{3},{4},{5},{6},{7},{8},{9}\n'.format(sub_row.iloc[0]['имя'],
                                                                        sub_row.iloc[0]['коммент'],
                                                                        sub_row.iloc[0]['долг'], num,
                                                                        sub_row.iloc[0]['дата'],
                                                                        sub_row.iloc[0]['телефон'],
                                                                        day, time, comm, j))
                fcv.close()
                format_excel(doc_loans, sheet_loans)
                adjust_excel(doc_loans, sheet_loans)
                pd.set_option('display.max_rows', None)
                pd.set_option('display.max_columns', None)
                dataframe_subs = pd.read_csv(path_subs_csv, engine='python', encoding=enc)
                dataframe_subs.columns = ['имя', 'коммент', 'долг', 'изменение', 'дата',
                                          'телефон', 'дата изм.', 'время изм.', 'комментарий к изм.', 'описание']
                print('\n')
                print(dataframe_subs.iloc[-1])

            elif choice == 3:
                pd.set_option('display.max_rows', None)
                pd.set_option('display.max_columns', None)
                dataframe_subs = pd.read_csv(path_subs_csv, engine='python', encoding=enc)
                dataframe_subs.columns = ['имя', 'коммент', 'долг', 'изменение', 'дата',
                                          'телефон', 'дата изм.', 'время изм.', 'комментарий к изм.', 'описание']
                print('\n')
                print(dataframe_subs)

        except ValueError:
            print('\n' + ' Ошибка! Введите цифровое значение')
        except IndexError:
            print('\n' + ' Ошибка! Такого индекса не существует')
        except PermissionError:
            print('\n' + ' Ошибка! Закройте excel файл')


def search_func():
    while True:
        df()
        search = input(' ВВЕДИТЕ имя | коммент | долг | телефон | дата | "дни" | "все": ')
        search1 = search[:1].upper() + search[1:].lower()
        search2 = search.lower()
        search_int = None if not str(search).isdigit() else int(search)

        if search == '0' or search == 'выйти' or search == 'назад':
            break

        if search == 'дни':
            try:
                n = int(input(' ВВЕДИТЕ количество дней назад: '))
                days = timedelta(days=n)
                today = datetime.strptime(datetime.today().strftime('%d.%m.%Y'), '%d.%m.%Y')
                dd = today - days
                df_copy4 = df().copy()
                df_copy4['дата'] = pd.to_datetime(df_copy4['дата'], dayfirst=True)
                res_df = df_copy4.loc[df_copy4['дата'] <= dd]

                print('\n')
                print(res_df)
                total = res_df.долг.sum()
                print(f' СУММАРНЫЙ ДОЛГ = {total} руб.')
                print('\n')

                '''fr3 = open(path_csv, 'r', encoding=enc)
                lines = fr3.readlines()
                count = 1
                total = 0
                print('\n')
                for line in lines[1:]:
                    lst = line.split(',')
                    if datetime.strptime(lst[3], '%d.%m.%Y').date() <= dd:
                        print(f'{count}) {lst[0]} - ({lst[1]}) - {lst[2]} руб.'
                              f' - {lst[3]} - {lst[4]} - {lst[5][0:len(lst[5]) - 1]}')
                        dbt = int(lst[2])
                        total += dbt
                        count += 1
                print(f' СУММАРНЫЙ ДОЛГ = {total} руб.')
                print('\n')
                fr3.close()'''
            except ValueError:
                print(' Ошибка! Попробуйте снова\n')
                continue
            except IndexError:
                print(' Ошибка! Попробуйте снова\n')
                continue
            except OverflowError:
                print(' Ошибка! Попробуйте снова\n')
                continue

        elif search == 'все':
            df()
            print('\n')
            print(df())
            total = df().долг.sum()
            print(f' СУММАРНЫЙ ДОЛГ = {total} руб.')
            print('\n')

        elif search != '0' or search != 'выйти' or search != 'назад' or search != 'все' or search != 'дни':
            df()
            res_dt = df().loc[(df().имя == search) | (df().имя == search1) | (df().имя == search2)
                              | (df().коммент == search) | (df().коммент == search2)
                              | (df().долг == search_int) | (df().дата == search) | (df().дата == search2)
                              | (df().телефон.apply(lambda x: str(x)) == search)]
            if res_dt.empty:
                print(' Не найдено! Попробуйте снова')
            else:
                print('\n')
                print(res_dt)
                total = res_dt.долг.sum()
                print(f' СУММАРНЫЙ ДОЛГ = {total} руб.')
                print('\n')

        df()

        continue
